#!/usr/bin/env python2.7
# -*- coding: utf-8 -*-
# 2023/07/15: add lasting_days

import re 
import sys 
import os
import urllib 
import requests
import random
import json
import csv
import codecs
from urllib import unquote
import time
reload(sys)
sys.setdefaultencoding("utf-8")

from openpyxl import Workbook
import urllib3
urllib3.disable_warnings()

ticket_headers = {
    'user-agent': "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/75.0.3770.142 Safari/537.36",
    }

# force the number in list to int or float
def convert_list_format(info_list):
    info_new = []
    for x in info_list:
        if re.search(r'(^[0-9]{1,5}(\.0)?$)', x):
            x = int(float(x))
        elif re.search(r'(^[0-9]{1,2}\.[0-9]{2}$)', x):
            x = float(x)
        else:
            x = x
        info_new.append(x)
    return info_new

# write to csv file
def write_to_csv(filename, head_line, *info_list):
    with open(filename, 'w') as f:
        f.write(codecs.BOM_UTF8)
        writer = csv.writer(f)
        writer.writerow(head_line.split('\t'))
        for row in info_list:
            row_list = row.split('\t')
            writer.writerow(row_list)

# write to excel file
#def write_to_excel(filename, *info_list):
def write_to_excel(*info_list):
    wb = Workbook(write_only=True)
    ws = []
    host_tag_lists = [u'排片表', u'排片表-旧']
    row_list_old = []
    # create sheet
    for i in range(len(host_tag_lists)):
        ws.append(wb.create_sheet(title=host_tag_lists[i]))  # utf8->unicode
    # insert sheet header
    ws[0].append(['film', 'date', 'time', 'week', 'duration', 'year', 'fare', 'theater', 'movieHall', 'country', 'director', 'language', 'subtitle', 'projection_material', 'frameRatio'])
    #ws[1].append(['日期', '星期', '放映时间', '片名', '国家', '导演', '时长', '影院', '影厅', '票价'])
    ws[1].append(['日期', '星期', '放映时间', '影片信息', '时长', '影院', '影厅', '票价'])
    for row in info_list:
        row_info = row.split('\t')
        row_list = convert_list_format(row_info[0:5] \
                                      + row_info[15:13:-1] + row_info[5:7] \
                                      + row_info[8:6:-1] + row_info[9:13])
        theater = row_info[5].replace('艺术影院', '')
        movie_info = row_info[0] + '\n' + row_info[15] + '\n' \
                     + row_info[8] + ' | ' + row_info[7]
        #row_list_old.append([row_info[1], row_info[3], row_info[13], row_info[0], row_info[8], row_info[7], row_info[4], theater, row_info[6], row_info[14]])
        row_list_old.append([row_info[1], row_info[3], row_info[13], movie_info, row_info[4], theater, row_info[6], row_info[14]])
        ws[0].append(row_list)
    #for row in sorted(row_list_old, key=lambda x:(x[7], x[0], x[2]), reverse=False):
    # x[5]--theater, x[0]--row_info[1]--date, x[2]--row_info[13]--playTime
    for row in sorted(row_list_old, key=lambda x:(x[5], x[0], x[2]), reverse=False):
        row_list = convert_list_format(row)
        ws[1].append(row_list)
    # define the filename and save it to local disk
    save_path = 'zlg_schedule'
    save_path += '.xlsx'
    wb.save(save_path)


def _to_timestamp(dt):
    timeArray = time.strptime(dt, "%Y-%m-%d %H:%M:%S")
    ts = time.mktime(timeArray)
    return ts


def _to_dt(ts):
    timeArray = time.localtime(ts)
    dt = time.strftime("%H:%M", timeArray)
    return dt


def _to_day(ts):
    timeArray = time.localtime(ts)
    day = time.strftime("%Y/%m/%d", timeArray)
    return day


def get_week_day(dt):
    week_day_dict = {
      0 : '周日',
      1 : '周一',
      2 : '周二',
      3 : '周三',
      4 : '周四',
      5 : '周五',
      6 : '周六',
    }
    timeArray = time.strptime(dt, "%Y-%m-%d %H:%M:%S")
    day = int(time.strftime("%w", timeArray))
    return week_day_dict[day]


def get_schedule_info(date):
    #millis = int(round(time.time() * 1000))
    year = date.split('/')[0]
    month = date.split('/')[1]
    cinema_url = 'http://api.guoyingjiaying.cn/filmcinema/getcalendar?year={0}&month={1}&cinema_code=' .format(year, month)
    print "cinema_url:{}" .format(cinema_url)
    try:
        res = requests.get(cinema_url, timeout=5, headers=ticket_headers, verify=False)
        json_data = [ j for x in res.json()['data']['list'] if len(x['screen']) >0 for j in x['screen'] ]
    except requests.exceptions.RequestException as e:
        json_data = []
    print json_data
    return json_data

#sys.exit(0)


def get_movie_info(m_id):
    movie_info = {}
    movie_url = 'http://api.guoyingjiaying.cn/filmcinema/getprogram_details_app?prorgam_id={0}' .format(m_id)  
    res = requests.get(movie_url, headers=ticket_headers, verify=False)
    json_data=res.json()['data']
    movieActorList = json_data['performer']
    print "movieActorList:{0}" .format(movieActorList)
    # get all director
    movie_info['director_all'] = 'N/A' if len(movieActorList) == 0 else "/" .join([ x['real_name'].split()[0].strip() for x in movieActorList \
               if x['position'] == '导演'.decode('utf-8') ])
    # 国家/地区
    movie_info['regionCategoryName'] = "/" .join([ x for x in json_data['film_area_code'].split() ])
    # 语言
    movie_info['languageCategoryName'] = "/" .join([ x for x in json_data['film_language_code'].split() ])
    # 字幕
    movie_info['subtitle'] = '未标注' if len(json_data['film_caption_code']) == 0 else "/" .join([ x for x in json_data['film_caption_code'].split() ])
    # 版本 DCP
    #movie_info['projection_material'] = 'N/A' if len(json_data['film_pformat_code']) == 0 else json_data['film_pformat_code']
    movie_info['projection_material'] = 'N/A' if json_data['screens'][0]['screen_format'] == '' else json_data['screens'][0]['screen_format']
    # 画幅比
    try:
        movie_info['frameRatio'] = 'N/A' if json_data['film_frameratio_code'] == '' else json_data['film_frameratio_code']
    except Exception:
        movie_info['frameRatio'] = 'N/A'
    return movie_info


def get_detailed_schedule_info(schedule_data, cinema_list):
    for movie in schedule_data:
        if movie['screen_cinema'] in cinema_list:
            movie_id = movie['movie_id']
            program_id = movie['program_ids']
            print "program_id:{}" .format(program_id)
            name = movie['show_name']
            screen_cinema = movie['screen_cinema']
            cinema_name = screen_cinema.split()[0]
            movieHall = screen_cinema.split()[1]
            duration = movie['screen_time_len']
            fare = int(float(movie['show_price']))
            year = movie['film_year']
            poster = movie['cover_img1']
            movie_data = get_movie_info(program_id)
            # get the first three director
            director = "/" .join(movie_data['director_all'].split('/')[0:3])
            # get the first three country
            country = "/" .join(movie_data['regionCategoryName'].split('/')[0:3])
            print "country:{0}" .format(country)
            # get the first three language
            language = "/" .join(movie_data['languageCategoryName'].split('/')[0:3])
            #print "language:{0}" .format(language)
            # get date, beginTime, endTime, week
            playTime = movie['screen_start_time']
            showDate_list = movie['screen_start_time'].split()[0].split('-')
            showDate = "{0}月{1}日" .format(showDate_list[1], showDate_list[2])
            beginTime = ":" .join(movie['screen_start_time'].split()[1].split(':')[0:2])
            ts_start = _to_timestamp(playTime)
            ts_endTime = duration * 60 + ts_start
            endTime = _to_dt(ts_endTime)
            week = get_week_day(playTime)
            #movie_info = "{0}\t{1}\t{2}-{3}\t{4}\t{5}\t{6}\t{7}\t{8}" .format(name,showDate,beginTime,endTime,cinema_name,director,shot_year,country,poster)
            movie_info = "{0}\t{1}\t{2}-{3}\t{4}\t{5}\t{6}\t{7}\t{8}\t{9}\t{10}\t{11}\t{12}\t{13}\t{14}\t{15}\t{16}" \
                         .format(name,showDate,beginTime,endTime,week,duration,
                                 cinema_name,movieHall,director,country,language,
                                 movie_data['subtitle'],movie_data['projection_material'],
                                 movie_data['frameRatio'],beginTime,fare,year)
            movie_info_list.append(movie_info)
            time.sleep(0 + random.randint(0, 1000)/1000) 
            print cinema_name,duration,name,movieHall,poster,director
            print movie_info


def judge_list_dup_element(list_name, n, v=0):
    # check if continuous n elements in a list
    # `n` stand for the number of continuous element
    # `v` stand for the value of dup element
    for i in xrange(len(list_name)-n+1):
        if list_name[i] == v and len(list(set(list_name[i:i+n]))) == 1:
            return True
            break
    return False


def get_movie_detailed_info(start_day, cinema='北京总馆', lasting_days=31):
    if cinema == '北京总馆':
        cinema_list = ["小西天艺术影院 2号厅", "小西天艺术影院 1号厅", "百子湾艺术影院 1号厅"]
    elif cinema == '江南分馆':
        cinema_list = ["江南分馆影院 1号厅", "江南分馆影院 4号厅"]
    else:
        cinema_list = ["小西天艺术影院 2号厅", "小西天艺术影院 1号厅", "百子湾艺术影院 1号厅", "江南分馆影院 1号厅", "江南分馆影院 4号厅"]
    ts_start_day = _to_timestamp(start_day)
    ts_end_day = ts_start_day + lasting_days*24*60*60
    month_list = []
    while ts_start_day <= ts_end_day:
        month  =  "/" .join(_to_day(ts_start_day).split('/')[0:2])
        if month not in month_list:
            schedule_data = get_schedule_info(month)
            if len(schedule_data) > 0:
                get_detailed_schedule_info(schedule_data, cinema_list)
        month_list.insert(0, month)
        ts_start_day += 86400
    return movie_info_list


def get_schedule_list():
    # execute on the last 6 days of every month to get the movie schedule of next month
    ts_today = int(time.time())
    for i in xrange(6, 0, -1):
        # timestamp of tomorrow
        ts = ts_today + i * 86400
        year = _to_day(ts).split('/')[0]
        mon = _to_day(ts).split('/')[1]
        day = _to_day(ts).split('/')[-1]
        day_num = int(day)
        if day_num == 1:
            s_day = "{0}-{1}-{2} 00:00:00" .format(year, mon, day)
            movie_info_list = get_movie_detailed_info(s_day)
            # add file suffix
            write_to_csv(f_csv+'.'+year+mon, head_instruction, *movie_info_list)


if __name__ == '__main__':
    movie_info_list = []
    # write to movie.csv
    BASEPATH = os.path.realpath(os.path.dirname(__file__))
    f_csv = BASEPATH + os.sep + 'movie.csv'
    head_instruction = "film\tdate\ttime\tweek\tduration\ttheater\tmovieHall\tdirector\tcountry\tlanguage\tsubtitle\tprojection_material\tframeRatio\tplayTime\tfare\tyear"
    start_day = "2024-08-01 00:00:00"
    if len(sys.argv) > 1:
        if sys.argv[1] == 'all':
            cinema_name = 'all'
        else:
            cinema_name = '江南分馆'
    else:
        cinema_name = '北京总馆'
    movie_info_list = get_movie_detailed_info(start_day, cinema_name, 31) # lasting_days
    write_to_csv(f_csv, head_instruction, *movie_info_list)
    write_to_excel(*movie_info_list)
    sys.exit(0)
    get_schedule_list()
